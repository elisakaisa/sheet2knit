from openpyxl import load_workbook
from .pattern import Pattern


def excel_rgb_to_hex(rgb):
    """Convert an Excel ARGB colour (e.g. 'FFFF0000') to '#ff0000'."""
    if rgb is None:
        return None

    return "#" + rgb[-6:].lower()

def crop_pattern(rows):
    """Remove empty border rows and columns."""

    if not rows:
        return rows

    # Remove empty rows at the top
    while rows and all(cell is None for cell in rows[0]):
        rows.pop(0)

    # Remove empty rows at the bottom
    while rows and all(cell is None for cell in rows[-1]):
        rows.pop()

    if not rows:
        return rows

    # Find leftmost and rightmost non-empty columns
    left = min(
        i
        for row in rows
        for i, cell in enumerate(row)
        if cell is not None
    )

    right = max(
        i
        for row in rows
        for i, cell in enumerate(row)
        if cell is not None
    )

    return [row[left:right + 1] for row in rows]

def read_pattern(filename):
    wb = load_workbook(filename)
    ws = wb.active

    rows = []

    for row in ws.iter_rows():
        current_row = []

        for cell in row:
            fill = cell.fill

            if fill.fill_type is None:
                current_row.append(None)
            else:
                current_row.append(excel_rgb_to_hex(fill.fgColor.rgb))

        rows.append(current_row)

    rows = crop_pattern(rows)
    return Pattern(rows)